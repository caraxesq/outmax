from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Any

from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.db.models import Recipient


@dataclass
class ImportResult:
    imported: int
    skipped: int
    duplicates: int
    invalid: int


class RecipientImporter:
    def __init__(self, sessionmaker: async_sessionmaker[AsyncSession]):
        self.sessionmaker = sessionmaker

    async def import_csv_text(self, csv_text: str) -> ImportResult:
        reader = csv.DictReader(io.StringIO(csv_text))
        imported = skipped = duplicates = invalid = 0
        seen: set[str] = set()

        async with self.sessionmaker() as session:
            for row in reader:
                normalized = {str(k).strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
                user_id = self._parse_user_id(normalized.get("user_id"))
                username = self._normalize_username(normalized.get("username"))
                if user_id is None and username is None:
                    invalid += 1
                    continue
                key = f"id:{user_id}" if user_id is not None else f"username:{username}"
                if key in seen:
                    duplicates += 1
                    continue
                seen.add(key)
                exists = await session.scalar(
                    select(Recipient).where(
                        or_(
                            Recipient.user_id == user_id if user_id is not None else False,
                            Recipient.username == username if username is not None else False,
                        )
                    )
                )
                if exists:
                    duplicates += 1
                    continue
                do_not_contact = str(normalized.get("do_not_contact", "")).lower() in {"1", "true", "yes", "y"}
                metadata = self._metadata(normalized)
                session.add(
                    Recipient(
                        user_id=user_id,
                        username=username,
                        segment=normalized.get("segment") or None,
                        metadata_json=metadata,
                        do_not_contact=do_not_contact,
                    )
                )
                imported += 1
            await session.commit()
        return ImportResult(imported=imported, skipped=skipped, duplicates=duplicates, invalid=invalid)

    async def import_csv_bytes(self, data: bytes) -> ImportResult:
        return await self.import_csv_text(data.decode("utf-8-sig"))

    async def import_usernames_text(self, text: str) -> ImportResult:
        usernames = self.extract_usernames(text)
        return await self.import_usernames(usernames)

    async def import_table_bytes(self, data: bytes, file_name: str | None = None) -> ImportResult:
        name = (file_name or "").lower()
        if name.endswith(".xlsx"):
            return await self.import_usernames(self._extract_xlsx_usernames(data))
        try:
            return await self.import_usernames_text(data.decode("utf-8-sig"))
        except UnicodeDecodeError:
            return await self.import_usernames_text(data.decode("cp1251", errors="ignore"))

    async def import_usernames(self, usernames: list[str]) -> ImportResult:
        imported = duplicates = invalid = 0
        seen: set[str] = set()
        async with self.sessionmaker() as session:
            for username in usernames:
                normalized = self._normalize_username(username)
                if normalized is None:
                    invalid += 1
                    continue
                if normalized in seen:
                    duplicates += 1
                    continue
                seen.add(normalized)
                exists = await session.scalar(select(Recipient).where(Recipient.username == normalized))
                if exists:
                    duplicates += 1
                    continue
                session.add(Recipient(username=normalized, metadata_json={"username": normalized}))
                imported += 1
            await session.commit()
        return ImportResult(imported=imported, skipped=0, duplicates=duplicates, invalid=invalid)

    @classmethod
    def extract_usernames(cls, text: str) -> list[str]:
        usernames: list[str] = []
        for match in re.finditer(r"(?:https?://t\.me/|t\.me/|@)?([A-Za-z][A-Za-z0-9_]{2,31})", text):
            candidate = match.group(1)
            if candidate.lower() in {"http", "https", "user_id", "username", "name", "segment"}:
                continue
            usernames.append(candidate)
        return usernames

    def _extract_xlsx_usernames(self, data: bytes) -> list[str]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        usernames: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is not None:
                        usernames.extend(self.extract_usernames(str(value)))
        workbook.close()
        return usernames

    @staticmethod
    def _parse_user_id(value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return int(str(value))
        except ValueError:
            return None

    @staticmethod
    def _normalize_username(value: Any) -> str | None:
        if value in (None, ""):
            return None
        username = str(value).strip()
        if username.startswith("@"):
            username = username[1:]
        return username.lower() or None

    @staticmethod
    def _metadata(row: dict[str, Any]) -> dict[str, Any]:
        excluded = {"user_id", "username", "segment", "do_not_contact"}
        return {key: value for key, value in row.items() if key not in excluded and value not in (None, "")}
